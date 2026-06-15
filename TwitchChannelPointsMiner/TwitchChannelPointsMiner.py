# -*- coding: utf-8 -*-

from __future__ import annotations

import logging
import os
import random
import signal
import shutil
import sys
import threading
import time
import uuid
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from TwitchChannelPointsMiner.classes.Chat import ChatPresence, ThreadChat
from TwitchChannelPointsMiner.classes.PubSub import PubSubHandler
from TwitchChannelPointsMiner.classes.entities.PubsubTopic import PubsubTopic
from TwitchChannelPointsMiner.classes.entities.Streamer import (
    Streamer,
    StreamerSettings,
)
from TwitchChannelPointsMiner.classes.Exceptions import StreamerDoesNotExistException
from TwitchChannelPointsMiner.classes.Settings import FollowersOrder, Priority, Settings
from TwitchChannelPointsMiner.classes.Twitch import Twitch
from TwitchChannelPointsMiner.classes.websocket import (
    HermesWebSocketPool,
    PubSubWebSocketPool,
)
from TwitchChannelPointsMiner.classes.websocket.hermes.data import JsonDecoder, JsonEncoder
from TwitchChannelPointsMiner.constants import (
    CLIENT_ID_WEB,
    FORK_OWNER,
    GITHUB_REPO_URL,
    HERMES_WEBSOCKET,
)
from TwitchChannelPointsMiner.logger import LoggerSettings, configure_loggers
from TwitchChannelPointsMiner.WatchStreakCache import (
    MIN_OFFLINE_FOR_NEW_STREAK,
    WatchStreakCache,
)
from TwitchChannelPointsMiner.utils import (
    _millify,
    at_least_one_value_in_settings_is,
    check_versions,
    dump_json,
    get_user_agent,
    internet_connection_available,
    interruptible_sleep,
    load_json,
    set_default_settings,
)

# Suppress:
#   - chardet.charsetprober - [feed]
#   - chardet.charsetprober - [get_confidence]
#   - requests - [Starting new HTTPS connection (1)]
#   - Flask (werkzeug) logs
#   - irc.client - [process_data]
#   - irc.client - [_dispatcher]
#   - irc.client - [_handle_message]
logging.getLogger("chardet.charsetprober").setLevel(logging.ERROR)
logging.getLogger("requests").setLevel(logging.ERROR)
logging.getLogger("werkzeug").setLevel(logging.ERROR)
logging.getLogger("irc.client").setLevel(logging.ERROR)
logging.getLogger("seleniumwire").setLevel(logging.ERROR)
logging.getLogger("websocket").setLevel(logging.ERROR)

logger = logging.getLogger(__name__)
REPORT_PERIODS = ("daily", "weekly", "monthly", "yearly")
REPORT_POINT_COLUMNS = {
    "daily": "Daily Points",
    "weekly": "Weekly Points",
    "monthly": "Monthly Points",
    "yearly": "Yearly Points",
}
REPORT_METADATA_COLUMNS = [
    "Streamer",
    "Points",
    "Followdate",
    "Last Stream",
    "Sub",
    "Banned",
    "Watchstreaks",
]
REPORT_COLUMNS_BY_PERIOD = {
    "daily": REPORT_METADATA_COLUMNS
    + [
        REPORT_POINT_COLUMNS["daily"],
        REPORT_POINT_COLUMNS["weekly"],
        REPORT_POINT_COLUMNS["monthly"],
        REPORT_POINT_COLUMNS["yearly"],
    ],
    "weekly": REPORT_METADATA_COLUMNS
    + [
        REPORT_POINT_COLUMNS["weekly"],
        REPORT_POINT_COLUMNS["monthly"],
        REPORT_POINT_COLUMNS["yearly"],
    ],
    "monthly": REPORT_METADATA_COLUMNS
    + [
        REPORT_POINT_COLUMNS["monthly"],
        REPORT_POINT_COLUMNS["yearly"],
    ],
    "yearly": REPORT_METADATA_COLUMNS + [REPORT_POINT_COLUMNS["yearly"]],
}


class TwitchChannelPointsMiner:
    __slots__ = [
        "username",
        "twitch",
        "claim_drops_startup",
        "enable_analytics",
        "disable_ssl_cert_verification",
        "disable_at_in_nickname",
        "use_hermes",
        "priority",
        "streamers",
        "events_predictions",
        "minute_watcher_thread",
        "sync_campaigns_thread",
        "ws_pool",
        "session_id",
        "running",
        "start_datetime",
        "original_streamers",
        "logs_file",
        "queue_listener",
        "watch_streak_cache_path",
        "watch_streak_cache",
        "watch_streak_max_parallel",
        "watch_streak_min_offline_seconds",
        "streamers_export_path",
        "streamers_export_thread",
        "streamers_export_interval_seconds",
        "daily_reports",
        "weekly_reports",
        "monthly_reports",
        "yearly_reports",
        "streamer_follow_dates",
        "daily_points_baseline_path",
        "subscription_notification_cache_path",
        "daily_points_day_key",
        "daily_points_baseline",
        "daily_points_snapshot",
        "daily_points_session_anchor",
        "_daily_points_baseline_dirty",
        "period_points_baseline",
        "period_points_snapshot",
        "period_points_session_anchor",
        "period_points_keys",
        "_period_points_baseline_dirty",
        "_watch_streak_days_lookup_attempted",
        "_chat_ban_lookup_attempted",
    ]

    def __init__(
        self,
        username: str,
        password: str = None,
        claim_drops_startup: bool = False,
        enable_analytics: bool = False,
        disable_ssl_cert_verification: bool = False,
        disable_at_in_nickname: bool = False,
        # Settings for logging and selenium as you can see.
        priority: list[Priority] | Priority | None = None,
        # This settings will be global shared trought Settings class
        logger_settings: LoggerSettings = LoggerSettings(),
        # Default values for all streamers
        streamer_settings: StreamerSettings = StreamerSettings(),
        watch_streak_max_parallel: int | None = None,
        watch_streak_min_offline_seconds: int = MIN_OFFLINE_FOR_NEW_STREAK,
        use_hermes: bool = True,
        daily_reports: bool = True,
        weekly_reports: bool = False,
        monthly_reports: bool = False,
        yearly_reports: bool = False,
    ):
        # Fixes TypeError: 'NoneType' object is not subscriptable
        if not username or username == "your-twitch-username":
            logger.error("Please edit your runner file (usually run.py) and try again.")
            logger.error("No username, exiting...")
            sys.exit(0)

        # This disables certificate verification and allows the connection to proceed, but also makes it vulnerable to man-in-the-middle (MITM) attacks.
        Settings.disable_ssl_cert_verification = disable_ssl_cert_verification

        Settings.disable_at_in_nickname = disable_at_in_nickname
        Settings.use_hermes = use_hermes

        # Wait for Twitch.tv connectivity with a timeout to avoid hanging forever
        error_printed = False
        connectivity_interval = 5
        connectivity_timeout = 60
        connectivity_start = time.time()
        while not internet_connection_available(host="twitch.tv", port=443):
            if not error_printed:
                logger.error("Waiting for Twitch.tv connectivity...")
                error_printed = True
            if (time.time() - connectivity_start) >= connectivity_timeout:
                logger.error(
                    "Unable to reach Twitch.tv after 60 seconds, exiting..."
                )
                sys.exit(0)
            time.sleep(connectivity_interval)

        # Analytics switch
        Settings.enable_analytics = enable_analytics

        if enable_analytics is True:
            Settings.analytics_path = os.path.join(
                Path().absolute(), "analytics", username
            )
            Path(Settings.analytics_path).mkdir(parents=True, exist_ok=True)

        self.username = username

        # Set as global config
        Settings.logger = logger_settings

        # Init as default all the missing values
        streamer_settings.default()
        streamer_settings.bet.default()
        Settings.streamer_settings = streamer_settings

        # user_agent = get_user_agent("FIREFOX")
        user_agent = get_user_agent("CHROME")
        self.watch_streak_max_parallel = (
            max(1, int(watch_streak_max_parallel))
            if watch_streak_max_parallel is not None
            else None
        )
        self.watch_streak_min_offline_seconds = max(
            0, int(watch_streak_min_offline_seconds)
        )
        self.twitch = Twitch(
            self.username, user_agent, password, self.watch_streak_max_parallel
        )

        self.claim_drops_startup = claim_drops_startup
        self.use_hermes = use_hermes
        safe_account_name = "".join(
            ch if (ch.isalnum() or ch in "._-") else "_"
            for ch in self.username.lower()
        )
        self.watch_streak_cache_path = os.path.join(
            "logs", ".state", f"watch_streak_cache.{safe_account_name}.json"
        )
        report_date = datetime.now().strftime("%Y-%m-%d")
        self.streamers_export_path = os.path.join(
            "logs",
            "reports",
            "daily",
            f"report_{report_date}_{safe_account_name}.xlsx",
        )
        self.streamers_export_thread = None
        self.streamers_export_interval_seconds = 10 * 60
        self.daily_reports = bool(daily_reports)
        self.weekly_reports = bool(weekly_reports)
        self.monthly_reports = bool(monthly_reports)
        self.yearly_reports = bool(yearly_reports)
        self.streamer_follow_dates = {}
        self.daily_points_baseline_path = os.path.join(
            "logs", ".state", f"daily_points_baseline.{safe_account_name}.json"
        )
        self.subscription_notification_cache_path = os.path.join(
            "logs", ".state", f"subscription_notifications.{safe_account_name}.json"
        )
        self.daily_points_day_key = datetime.now().strftime("%Y-%m-%d")
        self.daily_points_baseline = {}
        self.daily_points_snapshot = {}
        self.daily_points_session_anchor = {}
        self._daily_points_baseline_dirty = False
        self.period_points_baseline = {}
        self.period_points_snapshot = {}
        self.period_points_session_anchor = {}
        self.period_points_keys = {}
        self._period_points_baseline_dirty = False
        self._watch_streak_days_lookup_attempted = set()
        self._chat_ban_lookup_attempted = set()
        self._load_daily_points_baseline()
        initial_cache_load_path = self._watch_streak_cache_load_path()
        self.watch_streak_cache = WatchStreakCache.load_from_disk(
            initial_cache_load_path,
            default_account_name=self.username,
            account_filter=self.username,
            min_offline_for_new_streak=self.watch_streak_min_offline_seconds,
        )
        self.twitch.watch_streak_cache = self.watch_streak_cache
        self.twitch.subscription_notification_cache_path = (
            self.subscription_notification_cache_path
        )
        if priority is None:
            self.priority = [Priority.STREAK, Priority.DROPS, Priority.ORDER]
        elif isinstance(priority, Priority):
            self.priority = [priority]
        else:
            self.priority = priority

        self.streamers: list[Streamer] = []
        self.events_predictions = {}
        self.minute_watcher_thread = None
        self.sync_campaigns_thread = None
        self.ws_pool = None

        self.session_id = str(uuid.uuid4())
        self.running = False
        self.start_datetime = None
        self.original_streamers = []

        self.logs_file, self.queue_listener = configure_loggers(
            self.username, logger_settings
        )

        # Check for the latest version of the script
        current_version, github_version = check_versions()

        logger.info(
            f"Twitch Channel Points Miner v{current_version} (fork by {FORK_OWNER})"
        )
        logger.info(GITHUB_REPO_URL)

        if github_version == "0.0.0":
            logger.error(
                "Unable to detect if you have the latest version of this script"
            )
        elif current_version != github_version:
            logger.info(f"You are running version {current_version} of this script")
            logger.info(f"The latest version on GitHub is {github_version}")

        for sign in [signal.SIGINT, signal.SIGSEGV, signal.SIGTERM]:
            signal.signal(sign, self.end)

    def analytics(
        self,
        host: str = "127.0.0.1",
        port: int = 5000,
        refresh: int = 5,
        days_ago: int = 7,
    ):
        # Analytics switch
        if Settings.enable_analytics is True:
            from TwitchChannelPointsMiner.classes.AnalyticsServer import AnalyticsServer

            days_ago = days_ago if days_ago <= 365 * 15 else 365 * 15
            http_server = AnalyticsServer(
                host=host,
                port=port,
                refresh=refresh,
                days_ago=days_ago,
                username=self.username,
            )
            http_server.daemon = True
            http_server.name = "Analytics Thread"
            http_server.start()
        else:
            logger.error("Can't start analytics(), please set enable_analytics=True")

    def _format_followdate(self, followed_at: str | None) -> str:
        if not followed_at:
            return "..."
        try:
            normalized = str(followed_at).replace("Z", "+00:00")
            return datetime.fromisoformat(normalized).strftime("%d.%m.%Y")
        except Exception:
            return "..."

    def _format_sub(self, streamer: Streamer) -> str:
        return "yes" if streamer.subscription_tier is not None else "no"

    def _format_yes_no(self, value: object) -> str:
        return "yes" if bool(value) else "no"

    def _format_timestamp_date(self, timestamp: float | int | None) -> str:
        if timestamp in [None, ""]:
            return "..."
        try:
            return datetime.fromtimestamp(
                float(timestamp), tz=timezone.utc
            ).strftime("%d.%m.%Y")
        except Exception:
            return "..."

    def _safe_account_name(self) -> str:
        return "".join(
            ch if (ch.isalnum() or ch in "._-") else "_"
            for ch in self.username.lower()
        )

    def _legacy_state_path_for(self, path: str) -> str:
        state_dir = os.path.dirname(path)
        if os.path.basename(state_dir) == ".state":
            return os.path.join(os.path.dirname(state_dir), os.path.basename(path))
        return path

    def _first_existing_path(
        self,
        preferred_path: str,
        legacy_paths: list[str],
        label: str,
    ) -> str:
        if os.path.isfile(preferred_path):
            return preferred_path
        for legacy_path in legacy_paths:
            if legacy_path != preferred_path and os.path.isfile(legacy_path):
                try:
                    Path(os.path.dirname(preferred_path) or ".").mkdir(
                        parents=True,
                        exist_ok=True,
                    )
                    shutil.copy2(legacy_path, preferred_path)
                    logger.info(
                        "Copied legacy %s file to %s; kept original at %s",
                        label,
                        preferred_path,
                        legacy_path,
                    )
                    return preferred_path
                except Exception:
                    logger.warning(
                        "Failed to copy legacy %s file to %s; using %s for this startup",
                        label,
                        preferred_path,
                        legacy_path,
                    )
                    return legacy_path
        return preferred_path

    def _watch_streak_cache_load_path(self) -> str:
        safe_account_name = self._safe_account_name()
        return self._first_existing_path(
            self.watch_streak_cache_path,
            [
                self._legacy_state_path_for(self.watch_streak_cache_path),
                os.path.join("logs", "watch_streak_cache.json"),
                os.path.join("logs", f"watch_streak_cache.{safe_account_name}.json"),
            ],
            "watch streak cache",
        )

    def _daily_points_baseline_load_path(self) -> str | None:
        path = getattr(self, "daily_points_baseline_path", None)
        if not path:
            return path
        return self._first_existing_path(
            path,
            [self._legacy_state_path_for(path)],
            "daily points baseline",
        )

    def _uses_dated_streamers_export_path(self) -> bool:
        file_name = os.path.basename(self.streamers_export_path or "")
        return file_name.startswith("report_") and file_name.endswith(
            f"_{self._safe_account_name()}.xlsx"
        )

    def _current_streamers_export_path(self) -> str:
        output_dir = os.path.dirname(self.streamers_export_path) or "logs"
        report_date = datetime.now().strftime("%Y-%m-%d")
        return os.path.join(
            output_dir, f"report_{report_date}_{self._safe_account_name()}.xlsx"
        )

    def _report_output_dir(self, period: str) -> str:
        output_dir = os.path.dirname(self.streamers_export_path) or os.path.join(
            "logs", "reports", "daily"
        )
        reports_root = os.path.dirname(output_dir)
        if (
            os.path.basename(output_dir) == "daily"
            and os.path.basename(reports_root) == "reports"
        ):
            return os.path.join(reports_root, period)
        return output_dir

    def _period_report_path(self, period: str) -> str:
        output_dir = self._report_output_dir(period)
        account_name = self._safe_account_name()
        now = datetime.now()
        if period == "daily":
            return self._current_streamers_export_path()
        if period == "weekly":
            return os.path.join(
                output_dir,
                f"weekly_report_{self._period_key('weekly', now)}_{account_name}.xlsx",
            )
        if period == "monthly":
            return os.path.join(
                output_dir,
                f"monthly_report_{now.strftime('%B')}_{now.strftime('%Y')}_{account_name}.xlsx",
            )
        if period == "yearly":
            return os.path.join(
                output_dir,
                f"yearly_report_{now.strftime('%Y')}_{account_name}.xlsx",
            )
        raise ValueError(f"Unknown report period: {period}")

    def _enabled_report_paths(self) -> list[str]:
        enabled = {
            "daily": self.daily_reports,
            "weekly": self.weekly_reports,
            "monthly": self.monthly_reports,
            "yearly": self.yearly_reports,
        }
        return [
            self._period_report_path(period)
            for period in REPORT_PERIODS
            if enabled.get(period)
        ]

    def _enabled_report_targets(self) -> list[tuple[str, str]]:
        enabled = {
            "daily": self.daily_reports,
            "weekly": self.weekly_reports,
            "monthly": self.monthly_reports,
            "yearly": self.yearly_reports,
        }
        return [
            (period, self._period_report_path(period))
            for period in REPORT_PERIODS
            if enabled.get(period)
        ]

    def _refresh_streamers_export_path_if_needed(self) -> None:
        if not self._uses_dated_streamers_export_path():
            return
        current_path = self._current_streamers_export_path()
        if self.streamers_export_path != current_path:
            self.streamers_export_path = current_path

    def _last_stream_date(self, streamer: Streamer) -> str:
        created_at = getattr(getattr(streamer, "stream", None), "created_at", None)
        if created_at not in [None, ""]:
            return self._format_timestamp_date(created_at)

        watch_streak_cache = getattr(self, "watch_streak_cache", None)
        account_name = getattr(self, "username", None)
        if watch_streak_cache is None:
            return "..."

        try:
            status = watch_streak_cache.get_streamer_status(
                streamer.username,
                account_name=account_name,
            )
            candidate_timestamps: list[float] = []
            if (
                status is not None
                and getattr(status, "last_stream_started_at", None) not in [None, ""]
            ):
                candidate_timestamps.append(float(status.last_stream_started_at))

            latest_session = watch_streak_cache.latest_session_for_streamer(
                streamer.username,
                account_name=account_name,
            )
            if latest_session is not None and latest_session.started_at not in [None, ""]:
                candidate_timestamps.append(float(latest_session.started_at))

            if candidate_timestamps:
                return self._format_timestamp_date(max(candidate_timestamps))
        except Exception:
            return "..."
        return "..."

    def _history_points_total(self, streamer: Streamer) -> int:
        history = getattr(streamer, "history", {})
        if not isinstance(history, dict):
            return 0

        total = 0
        for entry in history.values():
            if not isinstance(entry, dict):
                continue
            amount = entry.get("amount", 0)
            if isinstance(amount, (int, float)):
                total += int(amount)
        return total

    def _period_key(self, period: str, now: datetime | None = None) -> str:
        current = now or datetime.now()
        if period == "daily":
            return current.strftime("%Y-%m-%d")
        if period == "weekly":
            iso_year, iso_week, _ = current.isocalendar()
            return f"{iso_year}-W{iso_week:02d}"
        if period == "monthly":
            return current.strftime("%Y-%m")
        if period == "yearly":
            return current.strftime("%Y")
        raise ValueError(f"Unknown report period: {period}")

    def _ensure_period_points_state(self) -> None:
        for period in REPORT_PERIODS:
            self.period_points_baseline.setdefault(period, {})
            self.period_points_snapshot.setdefault(period, {})
            self.period_points_session_anchor.setdefault(period, {})
            self.period_points_keys.setdefault(period, self._period_key(period))
        self.daily_points_day_key = self.period_points_keys["daily"]
        self.daily_points_baseline = self.period_points_baseline["daily"]
        self.daily_points_snapshot = self.period_points_snapshot["daily"]
        self.daily_points_session_anchor = self.period_points_session_anchor["daily"]

    def _reset_daily_points_baseline_if_needed(self) -> None:
        self._reset_period_points_baseline_if_needed()

    def _reset_period_points_baseline_if_needed(self) -> None:
        self._ensure_period_points_state()
        for period in REPORT_PERIODS:
            period_key = self._period_key(period)
            if self.period_points_keys.get(period) != period_key:
                self.period_points_keys[period] = period_key
                self.period_points_baseline[period] = {}
                self.period_points_snapshot[period] = {}
                self.period_points_session_anchor[period] = {}
                self._period_points_baseline_dirty = True
                if period == "daily":
                    self._daily_points_baseline_dirty = True
        self._ensure_period_points_state()

    def _load_daily_points_baseline(self) -> None:
        path = self._daily_points_baseline_load_path()
        self.period_points_baseline = {period: {} for period in REPORT_PERIODS}
        self.period_points_snapshot = {period: {} for period in REPORT_PERIODS}
        self.period_points_session_anchor = {period: {} for period in REPORT_PERIODS}
        self.period_points_keys = {
            period: self._period_key(period) for period in REPORT_PERIODS
        }
        self._period_points_baseline_dirty = False
        self._daily_points_baseline_dirty = False
        self._ensure_period_points_state()
        if not path:
            return

        payload = load_json(path, {})
        if not isinstance(payload, dict):
            return

        periods = payload.get("periods")
        if isinstance(periods, dict):
            for period in REPORT_PERIODS:
                bucket = periods.get(period)
                if not isinstance(bucket, dict):
                    continue
                if bucket.get("period_key") != self._period_key(period):
                    continue
                self.period_points_keys[period] = bucket["period_key"]
                points = bucket.get("points")
                if not isinstance(points, dict):
                    continue
                for username, value in points.items():
                    try:
                        self.period_points_baseline[period][str(username)] = int(value)
                    except Exception:
                        continue
                self.period_points_snapshot[period] = dict(
                    self.period_points_baseline[period]
                )
            self._ensure_period_points_state()
            return

        today = self._period_key("daily")
        if payload.get("day_key") != today:
            return

        stored_points = payload.get("points_gained")
        if not isinstance(stored_points, dict):
            legacy_baseline = payload.get("baseline")
            if isinstance(legacy_baseline, dict):
                logger.info(
                    "Ignoring legacy daily points baseline file %s; a new daily points snapshot will be written on the next export.",
                    path,
                )
            return
        if not isinstance(stored_points, dict):
            return

        normalized_baseline = {}
        for username, value in stored_points.items():
            try:
                normalized_baseline[str(username)] = int(value)
            except Exception:
                continue
        self.period_points_baseline["daily"] = normalized_baseline
        self.period_points_snapshot["daily"] = dict(normalized_baseline)
        self._ensure_period_points_state()

    def _save_daily_points_baseline_if_dirty(self) -> None:
        if not (
            getattr(self, "_daily_points_baseline_dirty", False)
            or getattr(self, "_period_points_baseline_dirty", False)
        ):
            return
        path = getattr(self, "daily_points_baseline_path", None)
        if not path:
            return
        self._ensure_period_points_state()

        dump_json(
            path,
            {
                "schema_version": 3,
                "day_key": self.period_points_keys["daily"],
                "points_gained": self.period_points_snapshot["daily"],
                "periods": {
                    period: {
                        "period_key": self.period_points_keys[period],
                        "points": self.period_points_snapshot[period],
                    }
                    for period in REPORT_PERIODS
                },
            },
        )
        self._daily_points_baseline_dirty = False
        self._period_points_baseline_dirty = False

    def _points_gained(self, streamer: Streamer) -> int:
        return self._points_for_period(streamer, "daily")

    def _points_for_period(self, streamer: Streamer, period: str) -> int:
        self._reset_period_points_baseline_if_needed()
        current_total = self._history_points_total(streamer)
        username = streamer.username
        carried_total = int(self.period_points_baseline[period].get(username, 0))
        anchor = self.period_points_session_anchor[period].get(username)
        if anchor is None:
            self.period_points_session_anchor[period][username] = current_total
            return carried_total

        session_delta = int(current_total - anchor)
        return int(carried_total + session_delta)

    def _sync_daily_points_snapshot(self, rows: list[dict[str, str]]) -> None:
        self._sync_period_points_snapshot(rows)

    def _sync_period_points_snapshot(self, rows: list[dict[str, str]]) -> None:
        self._ensure_period_points_state()
        changed = False
        for period in REPORT_PERIODS:
            column_name = REPORT_POINT_COLUMNS[period]
            snapshot = {}
            for row in rows:
                username = str(row.get("Streamer", "")).strip()
                if not username:
                    continue
                try:
                    snapshot[username] = int(row.get(column_name, 0))
                except Exception:
                    snapshot[username] = 0

            if snapshot != self.period_points_snapshot[period]:
                self.period_points_snapshot[period] = snapshot
                changed = True
                if period == "daily":
                    self._daily_points_baseline_dirty = True

        if changed:
            self._period_points_baseline_dirty = True
            self._ensure_period_points_state()

    def _fetch_watch_streak_days_from_twitch(self, streamer: Streamer) -> int | None:
        twitch = getattr(self, "twitch", None)
        if twitch is None or not getattr(streamer, "channel_id", None):
            return None

        attempted = getattr(self, "_watch_streak_days_lookup_attempted", None)
        if attempted is None:
            attempted = set()
            self._watch_streak_days_lookup_attempted = attempted
        if streamer.username in attempted:
            return None
        attempted.add(streamer.username)

        try:
            days = twitch.get_watch_streak_days(streamer)
        except Exception as exc:
            logger.debug(
                "Failed to fetch watch streak days from Twitch for %s: %s",
                streamer.username,
                exc,
            )
            return None
        if days in [None, ""]:
            return None

        days_value = max(0, int(days))
        watch_streak_cache = getattr(self, "watch_streak_cache", None)
        account_name = getattr(self, "username", None)
        if watch_streak_cache is not None:
            status = watch_streak_cache.get_streamer_status(
                streamer.username,
                account_name=account_name,
            )
            watch_streak_cache.set_streamer_status(
                streamer.username,
                watch_streak_detected=(
                    getattr(status, "watch_streak_detected", False)
                    if status is not None
                    else False
                ),
                is_online=(
                    getattr(status, "is_online", bool(getattr(streamer, "is_online", False)))
                    if status is not None
                    else bool(getattr(streamer, "is_online", False))
                ),
                watch_streak_days=days_value,
                last_stream_started_at=(
                    getattr(status, "last_stream_started_at", None)
                    if status is not None
                    else getattr(getattr(streamer, "stream", None), "created_at", None)
                ),
                broadcast_id=(
                    getattr(status, "broadcast_id", None)
                    if status is not None
                    else getattr(getattr(streamer, "stream", None), "broadcast_id", None)
                ),
                checked_at=time.time(),
                account_name=account_name,
            )
        return days_value

    def _fetch_chat_ban_status_from_twitch(self, streamer: Streamer) -> bool:
        twitch = getattr(self, "twitch", None)
        if twitch is None or not getattr(streamer, "channel_id", None):
            return bool(getattr(streamer, "chat_banned", False))

        attempted = getattr(self, "_chat_ban_lookup_attempted", None)
        if attempted is None:
            attempted = set()
            self._chat_ban_lookup_attempted = attempted
        if streamer.username in attempted:
            return bool(getattr(streamer, "chat_banned", False))
        attempted.add(streamer.username)

        try:
            chat_banned = twitch.get_chat_ban_status(streamer)
        except Exception as exc:
            logger.debug(
                "Failed to fetch chat ban status from Twitch for %s: %s",
                streamer.username,
                exc,
            )
            return bool(getattr(streamer, "chat_banned", False))
        if chat_banned is None:
            return bool(getattr(streamer, "chat_banned", False))

        streamer.chat_banned = bool(chat_banned)
        return streamer.chat_banned

    def _watch_streak_days(self, streamer: Streamer) -> int:
        watch_streak_cache = getattr(self, "watch_streak_cache", None)
        if watch_streak_cache is None:
            live_days = self._fetch_watch_streak_days_from_twitch(streamer)
            return live_days if live_days is not None else 0

        account_name = getattr(self, "username", None)
        try:
            status = watch_streak_cache.get_streamer_status(
                streamer.username,
                account_name=account_name,
            )
            if (
                status is not None
                and getattr(status, "watch_streak_days", None) not in [None, ""]
            ):
                return max(0, int(status.watch_streak_days))

            live_days = self._fetch_watch_streak_days_from_twitch(streamer)
            if live_days is not None:
                return live_days

            days = watch_streak_cache.claimed_streak_days(
                streamer.username,
                account_name=account_name,
            )
        except Exception:
            return 0
        return max(0, int(days))

    def _build_streamer_export_rows(self) -> list[dict[str, str]]:
        sorted_streamers = sorted(
            self.streamers,
            key=lambda streamer: (
                streamer.channel_points
                if isinstance(streamer.channel_points, (int, float))
                else float("-inf")
            ),
            reverse=True,
        )

        rows: list[dict[str, str]] = []
        for streamer in sorted_streamers:
            points_value = (
                streamer.channel_points
                if isinstance(streamer.channel_points, (int, float))
                else 0
            )
            rows.append(
                {
                    "Streamer": streamer.username,
                    "Points": _millify(points_value),
                    "Followdate": self._format_followdate(
                        self.streamer_follow_dates.get(streamer.username)
                    ),
                    "Last Stream": self._last_stream_date(streamer),
                    "Sub": self._format_sub(streamer),
                    "Banned": self._format_yes_no(
                        self._fetch_chat_ban_status_from_twitch(streamer)
                    ),
                    "Watchstreaks": self._watch_streak_days(streamer),
                    "Daily Points": self._points_for_period(streamer, "daily"),
                    "Weekly Points": self._points_for_period(streamer, "weekly"),
                    "Monthly Points": self._points_for_period(streamer, "monthly"),
                    "Yearly Points": self._points_for_period(streamer, "yearly"),
                }
            )
        self._sync_period_points_snapshot(rows)
        return rows

    def _write_streamers_xlsx(
        self,
        rows: list[dict[str, str]],
        output_path: str | None = None,
        period: str = "daily",
    ) -> None:
        if output_path is None:
            self._refresh_streamers_export_path_if_needed()
            output_path = self.streamers_export_path
        output_dir = os.path.dirname(output_path) or "."
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        columns = REPORT_COLUMNS_BY_PERIOD.get(period)
        if columns is None:
            raise ValueError(f"Unknown report period: {period}")

        data_frame = pd.DataFrame(
            rows,
            columns=columns,
        )

        tmp_fd, tmp_path = tempfile.mkstemp(
            prefix=".streamers_export_",
            suffix=".xlsx",
            dir=output_dir,
        )
        os.close(tmp_fd)
        try:
            data_frame.to_excel(tmp_path, index=False, engine="openpyxl")
            workbook = load_workbook(tmp_path)
            sheet = workbook.active
            self._style_streamers_export_sheet(sheet)

            workbook.save(tmp_path)
            os.replace(tmp_path, output_path)
        finally:
            if os.path.isfile(tmp_path):
                os.remove(tmp_path)

    def _style_streamers_export_sheet(self, sheet) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        success_fill = PatternFill("solid", fgColor="E8F5E9")
        soft_success_fill = PatternFill("solid", fgColor="F1F8E9")

        header_row = 1
        header_names = {}
        for col_idx in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=header_row, column=col_idx)
            header_name = str(header_cell.value or "")
            header_names[header_name] = col_idx
            header_cell.font = Font(bold=True, color="FFFFFF")
            header_cell.fill = header_fill
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[header_row].height = 22

        sheet.freeze_panes = "A2"

        for header_name in ("Followdate", "Last Stream"):
            col_idx = header_names.get(header_name)
            if col_idx is None:
                continue
            for row_idx in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value in [None, "", "..."]:
                    cell.value = None if cell.value == "..." else cell.value
                    continue
                if isinstance(cell.value, str):
                    try:
                        parsed = datetime.strptime(cell.value, "%d.%m.%Y").date()
                    except ValueError:
                        continue
                    cell.value = parsed
                cell.number_format = "DD.MM.YYYY"

        points_idx = header_names.get("Points")
        followdate_idx = header_names.get("Followdate")
        last_stream_idx = header_names.get("Last Stream")
        sub_idx = header_names.get("Sub")
        banned_idx = header_names.get("Banned")
        watch_idx = header_names.get("Watchstreaks")
        point_column_indices = [
            header_names[column_name]
            for column_name in REPORT_POINT_COLUMNS.values()
            if column_name in header_names
        ]
        max_positive_points = {col_idx: 0 for col_idx in point_column_indices}

        for row_idx in range(2, sheet.max_row + 1):
            if points_idx is not None:
                sheet.cell(row=row_idx, column=points_idx).alignment = Alignment(
                    horizontal="right", vertical="center"
                )
            for point_idx in point_column_indices:
                point_cell = sheet.cell(row=row_idx, column=point_idx)
                point_cell.alignment = Alignment(
                    horizontal="right", vertical="center"
                )
                if isinstance(point_cell.value, (int, float)):
                    max_positive_points[point_idx] = max(
                        max_positive_points[point_idx], int(point_cell.value)
                    )
                    if point_cell.value > 0:
                        point_cell.font = Font(color="2E7D32", bold=True)
                        point_cell.fill = soft_success_fill
                    elif point_cell.value < 0:
                        point_cell.font = Font(color="C62828", bold=True)
            if followdate_idx is not None:
                sheet.cell(row=row_idx, column=followdate_idx).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            if last_stream_idx is not None:
                sheet.cell(row=row_idx, column=last_stream_idx).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            if sub_idx is not None:
                sub_cell = sheet.cell(row=row_idx, column=sub_idx)
                sub_cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            if banned_idx is not None:
                banned_cell = sheet.cell(row=row_idx, column=banned_idx)
                banned_cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            if watch_idx is not None:
                watch_cell = sheet.cell(row=row_idx, column=watch_idx)
                watch_cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                if isinstance(watch_cell.value, (int, float)) and watch_cell.value >= 3:
                    watch_cell.font = Font(color="2E7D32", bold=True)
                    watch_cell.fill = success_fill
                elif isinstance(watch_cell.value, (int, float)) and watch_cell.value > 0:
                    watch_cell.font = Font(color="2E7D32")

        data_ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

        if sheet.max_row >= 2:
            table = Table(displayName="StreamersExportTable", ref=data_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)

            for point_idx, max_points in max_positive_points.items():
                if max_points <= 0:
                    continue
                point_column_letter = get_column_letter(point_idx)
                point_range = f"{point_column_letter}2:{point_column_letter}{sheet.max_row}"
                sheet.conditional_formatting.add(
                    point_range,
                    DataBarRule(
                        start_type="num",
                        start_value=0,
                        end_type="max",
                        end_value=0,
                        color="63C384",
                        showValue=True,
                    ),
                )

        column_widths = {
            "Streamer": 20,
            "Points": 10,
            "Followdate": 13,
            "Last Stream": 13,
            "Sub": 8,
            "Banned": 10,
            "Watchstreaks": 14,
            "Daily Points": 13,
            "Weekly Points": 14,
            "Monthly Points": 15,
            "Yearly Points": 14,
        }

        for col_idx in range(1, sheet.max_column + 1):
            header_name = str(sheet.cell(row=1, column=col_idx).value or "")
            width = column_widths.get(header_name)
            if width is not None:
                sheet.column_dimensions[get_column_letter(col_idx)].width = width

    def _export_streamers_snapshot(self) -> None:
        report_targets = self._enabled_report_targets()
        if not report_targets:
            return
        try:
            rows = self._build_streamer_export_rows()
            for period, report_path in report_targets:
                self._write_streamers_xlsx(rows, report_path, period=period)
        except Exception as exc:
            logger.warning(
                "Failed to update streamers export: %s",
                exc,
            )
        finally:
            self._save_daily_points_baseline_if_dirty()

    def _streamers_export_loop(self) -> None:
        while self.running:
            interruptible_sleep(
                lambda: self.running,
                self.streamers_export_interval_seconds,
            )
            if not self.running:
                break
            self._export_streamers_snapshot()

    def _create_ws_pool(self):
        listeners = [
            PubSubHandler(
                twitch=self.twitch,
                streamers=self.streamers,
                events_predictions=self.events_predictions,
            )
        ]
        if self.use_hermes:
            return HermesWebSocketPool(
                url=f"{HERMES_WEBSOCKET}?clientId={CLIENT_ID_WEB}",
                twitch=self.twitch,
                listeners=listeners,
                request_encoder=JsonEncoder(),
                response_decoder=JsonDecoder(),
            )
        return PubSubWebSocketPool(
            twitch=self.twitch,
            listeners=listeners,
        )

    def mine(
        self,
        streamers: list[Streamer | str] | None = None,
        blacklist: list[str] | None = None,
        followers: bool = False,
        followers_order: FollowersOrder = FollowersOrder.ASC,
    ):
        streamers_list = list(streamers) if streamers is not None else []
        blacklist_list = list(blacklist) if blacklist is not None else []
        self.run(
            streamers=streamers_list,
            blacklist=blacklist_list,
            followers=followers,
            followers_order=followers_order,
        )

    def run(
        self,
        streamers: list[Streamer | str] | None = None,
        blacklist: list[str] | None = None,
        followers: bool = False,
        followers_order: FollowersOrder = FollowersOrder.ASC,
    ):
        if self.running:
            logger.error("You can't start multiple sessions of this instance!")
            return

        streamers_input = list(streamers) if streamers is not None else []
        blacklist_input = list(blacklist) if blacklist is not None else []

        logger.info(f"Start session: '{self.session_id}'", extra={"emoji": ":bomb:"})
        self.running = True
        self.start_datetime = datetime.now()

        try:
            self.twitch.login()

            if self.claim_drops_startup is True:
                self.twitch.claim_all_drops_from_inventory()

            self.watch_streak_cache = WatchStreakCache.load_from_disk(
                self._watch_streak_cache_load_path(),
                default_account_name=self.username,
                account_filter=self.username,
                min_offline_for_new_streak=self.watch_streak_min_offline_seconds,
            )
            self.twitch.watch_streak_cache = self.watch_streak_cache
            self.twitch.subscription_notification_cache_path = (
                self.subscription_notification_cache_path
            )

            def normalize_login(name: str) -> str:
                return name.lower().strip().replace(" ", "")

            streamers_name: list = []
            streamers_dict: dict = {}

            for streamer in streamers_input:
                username = (
                    normalize_login(streamer.username)
                    if isinstance(streamer, Streamer)
                    else normalize_login(str(streamer))
                )
                if username not in blacklist_input:
                    streamers_name.append(username)
                    streamers_dict[username] = streamer

            if followers is True:
                followers_array = self.twitch.get_followers(order=followers_order)
                logger.info(
                    f"Load {len(followers_array)} followers from your profile!",
                    extra={"emoji": ":clipboard:"},
                )
                for username in followers_array:
                    if (
                        username not in streamers_dict
                        and normalize_login(username) not in blacklist_input
                    ):
                        norm = normalize_login(username)
                        streamers_name.append(norm)
                        streamers_dict[norm] = norm

            logger.info(
                f"Loading data for {len(streamers_name)} streamers. Please wait...",
                extra={"emoji": ":nerd_face:"},
            )
            load_workers = max(1, min(10, len(streamers_name))) if streamers_name else 0

            def build_streamer(username: str):
                streamer_obj = streamers_dict[username]
                streamer = (
                    streamer_obj
                    if isinstance(streamer_obj, Streamer) is True
                    else Streamer(username)
                )
                streamer.channel_id = self.twitch.get_channel_id(username)
                streamer.settings = set_default_settings(
                    streamer.settings, Settings.streamer_settings
                )
                streamer.settings.bet = set_default_settings(
                    streamer.settings.bet, Settings.streamer_settings.bet
                )
                if streamer.settings.chat != ChatPresence.NEVER:
                    streamer.irc_chat = ThreadChat(
                        self.username,
                        self.twitch.twitch_login.get_auth_token(),
                        streamer,
                    )
                streamer.watch_streak_cache = self.watch_streak_cache
                streamer.watch_streak_cache_path = self.watch_streak_cache_path
                streamer.watch_streak_account = self.username
                streamer.subscription_notification_cache_path = (
                    self.subscription_notification_cache_path
                )
                return streamer

            streamers_loaded = [None] * len(streamers_name)
            if streamers_name:
                with ThreadPoolExecutor(max_workers=load_workers or 1) as executor:
                    futures = {
                        executor.submit(build_streamer, username): index
                        for index, username in enumerate(streamers_name)
                    }
                    for future in as_completed(futures):
                        index = futures[future]
                        username = streamers_name[index]
                        try:
                            streamers_loaded[index] = future.result()
                        except StreamerDoesNotExistException:
                            logger.info(
                                f"Streamer {username} does not exist",
                                extra={"emoji": ":cry:"},
                            )
                        except Exception:
                            logger.error(
                                f"Failed to load streamer {username}", exc_info=True
                            )

            self.streamers = [
                streamer for streamer in streamers_loaded if streamer is not None
            ]

            # Populate the streamers with default values.
            # 1. Load channel points and auto-claim bonus
            # 2. Check if streamers are online
            # 3. DEACTIVATED: Check if the user is a moderator. (was used before the 5th of April 2021 to deactivate predictions)
            invalid_streamers = self.twitch.initialize_streamers_context(self.streamers)
            if invalid_streamers:
                self.streamers = [
                    streamer
                    for streamer in self.streamers
                    if streamer.username not in invalid_streamers
                ]
                if not self.streamers:
                    logger.error("No valid streamers available after initialization.")
                    self.end(0, 0)
                    return

            if self.watch_streak_cache is not None:
                snapshot_now = time.time()
                for streamer in self.streamers:
                    self.watch_streak_cache.set_streamer_status(
                        streamer.username,
                        watch_streak_detected=(
                            streamer.settings.watch_streak is True
                            and streamer.stream.watch_streak_missing is False
                        ),
                        is_online=bool(streamer.is_online),
                        last_stream_started_at=getattr(streamer.stream, "created_at", None),
                        broadcast_id=(
                            streamer.stream.broadcast_id if streamer.is_online else None
                        ),
                        checked_at=snapshot_now,
                        account_name=self.username,
                    )

                self.watch_streak_cache.mark_bootstrap_done()
                # Persist startup snapshot so long streamer lists can resume reliably after restart.
                self.watch_streak_cache.save_to_disk_if_dirty(
                    self.watch_streak_cache_path
                )

            self.original_streamers = [
                streamer.channel_points for streamer in self.streamers
            ]

            try:
                self.streamer_follow_dates = self.twitch.get_followers_with_dates(
                    order=followers_order
                )
            except Exception as exc:
                self.streamer_follow_dates = {}
                logger.warning("Failed to load follower dates for export: %s", exc)

            self._export_streamers_snapshot()
            self.streamers_export_thread = threading.Thread(
                target=self._streamers_export_loop
            )
            self.streamers_export_thread.name = "Streamers export"
            self.streamers_export_thread.start()

            # If we have at least one streamer with settings = make_predictions True
            make_predictions = at_least_one_value_in_settings_is(
                self.streamers, "make_predictions", True
            )

            # If we have at least one streamer with settings = claim_drops True
            # Spawn a thread for sync inventory and dashboard
            if (
                at_least_one_value_in_settings_is(self.streamers, "claim_drops", True)
                is True
            ):
                self.sync_campaigns_thread = threading.Thread(
                    target=self.twitch.sync_campaigns,
                    args=(self.streamers, 60),
                )
                self.sync_campaigns_thread.name = "Sync campaigns/inventory"
                self.sync_campaigns_thread.start()

            self.minute_watcher_thread = threading.Thread(
                target=self.twitch.send_minute_watched_events,
                args=(self.streamers, self.priority, 60),
            )
            self.minute_watcher_thread.name = "Minute watcher"
            self.minute_watcher_thread.start()

            self.ws_pool = self._create_ws_pool()
            self.ws_pool.start()

            # Subscribe to community-points-user. Get update for points spent or gains
            user_id = self.twitch.twitch_login.get_user_id()
            # print(f"!!!!!!!!!!!!!! USER_ID: {user_id}")

            # Fixes 'ERR_BADAUTH'
            if not user_id:
                logger.error("No user_id, exiting...")
                self.end(0, 0)

            self.ws_pool.submit(
                PubsubTopic(
                    "community-points-user-v1",
                    user_id=user_id,
                )
            )
            self.ws_pool.submit(
                PubsubTopic(
                    "user-subscribe-events-v1",
                    user_id=user_id,
                )
            )
            self.ws_pool.submit(
                PubsubTopic(
                    "onsite-notifications",
                    user_id=user_id,
                )
            )

            # Going to subscribe to predictions-user-v1. Get update when we place a new prediction (confirm)
            if make_predictions is True:
                self.ws_pool.submit(
                    PubsubTopic(
                        "predictions-user-v1",
                        user_id=user_id,
                    )
                )

            for streamer in self.streamers:
                self.ws_pool.submit(
                    PubsubTopic("video-playback-by-id", streamer=streamer)
                )

                if streamer.settings.follow_raid is True:
                    self.ws_pool.submit(PubsubTopic("raid", streamer=streamer))

                if streamer.settings.make_predictions is True:
                    self.ws_pool.submit(
                        PubsubTopic("predictions-channel-v1", streamer=streamer)
                    )

                if streamer.settings.claim_moments is True:
                    self.ws_pool.submit(
                        PubsubTopic("community-moments-channel-v1", streamer=streamer)
                    )

                if streamer.settings.community_goals is True:
                    self.ws_pool.submit(
                        PubsubTopic("community-points-channel-v1", streamer=streamer)
                    )

            refresh_context = time.time()

            while self.running:
                interruptible_sleep(lambda: self.running, random.uniform(20, 60))
                self.ws_pool.check_stale_connections()

                if ((time.time() - refresh_context) // 60) >= 30:
                    refresh_context = time.time()
                    for index in range(0, len(self.streamers)):
                        if self.streamers[index].is_online:
                            self.twitch.load_channel_points_context(
                                self.streamers[index]
                            )
        finally:
            self.running = False
            if self.streamers_export_thread is not None:
                self.streamers_export_thread.join()
            self._export_streamers_snapshot()
            if self.watch_streak_cache is not None:
                self.watch_streak_cache.save_to_disk_if_dirty(
                    self.watch_streak_cache_path
                )
            self._save_daily_points_baseline_if_dirty()

    def end(self, signum, frame):
        if not self.running:
            return
        
        logger.info("CTRL+C Detected! Please wait just a moment!")

        for streamer in self.streamers:
            if (
                streamer.irc_chat is not None
                and streamer.settings.chat != ChatPresence.NEVER
            ):
                streamer.leave_chat()
                if streamer.irc_chat.is_alive() is True:
                    streamer.irc_chat.join()

        self.running = self.twitch.running = False
        if self.ws_pool is not None:
            self.ws_pool.end()

        if self.minute_watcher_thread is not None:
            self.minute_watcher_thread.join()

        if self.sync_campaigns_thread is not None:
            self.sync_campaigns_thread.join()

        if self.streamers_export_thread is not None:
            self.streamers_export_thread.join()

        self._export_streamers_snapshot()

        # Check if all the mutex are unlocked.
        # Prevent breaks of .json file
        for streamer in self.streamers:
            if streamer.mutex.locked():
                streamer.mutex.acquire()
                streamer.mutex.release()

        self.__print_report()

        # Stop the queue listener to make sure all messages have been logged
        if self.watch_streak_cache is not None:
            self.watch_streak_cache.save_to_disk_if_dirty(
                self.watch_streak_cache_path
            )
        self._save_daily_points_baseline_if_dirty()
        self.queue_listener.stop()

        sys.exit(0)

    def __print_report(self):
        print("\n")
        logger.info(
            f"Ending session: '{self.session_id}'", extra={"emoji": ":stop_sign:"}
        )
        if self.logs_file is not None:
            logger.info(
                f"Logs file: {self.logs_file}", extra={"emoji": ":page_facing_up:"}
            )
        logger.info(
            f"Duration {datetime.now() - self.start_datetime}",
            extra={"emoji": ":hourglass:"},
        )

        if not Settings.logger.less and self.events_predictions != {}:
            print("")
            for event_id in self.events_predictions:
                event = self.events_predictions[event_id]
                if (
                    event.bet_confirmed is True
                    and event.streamer.settings.make_predictions is True
                ):
                    logger.info(
                        f"{event.streamer.settings.bet}",
                        extra={"emoji": ":wrench:"},
                    )
                    if event.streamer.settings.bet.filter_condition is not None:
                        logger.info(
                            f"{event.streamer.settings.bet.filter_condition}",
                            extra={"emoji": ":pushpin:"},
                        )
                    logger.info(
                        f"{event.print_recap()}",
                        extra={"emoji": ":bar_chart:"},
                    )

        print("")
        for streamer_index in range(0, len(self.streamers)):
            if self.streamers[streamer_index].history != {}:
                gained = (
                    self.streamers[streamer_index].channel_points
                    - self.original_streamers[streamer_index]
                )
                
                from colorama import Fore
                streamer_highlight = Fore.YELLOW
                
                streamer_gain = (
                    f"{streamer_highlight}{self.streamers[streamer_index]}{Fore.RESET}, Total Points Gained: {_millify(gained)}"
                    if Settings.logger.less
                    else f"{streamer_highlight}{repr(self.streamers[streamer_index])}{Fore.RESET}, Total Points Gained (after farming - before farming): {_millify(gained)}"
                )
                
                indent = ' ' * 25
                streamer_history = '\n'.join(f"{indent}{history}" for history in self.streamers[streamer_index].print_history().split('; ')) 
                
                logger.info(
                    f"{streamer_gain}\n{streamer_history}",
                    extra={"emoji": ":moneybag:"},
                )
